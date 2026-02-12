"""
Excel Writer
Writes test cases to Excel format compatible with the test case parser
"""
import logging
import os
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import TestCase

logger = logging.getLogger(__name__)

class ExcelWriter:
    """
    Writes test cases to Excel files in the required format
    """
    
    # Column headers matching the parser format
    HEADERS = [
        "Test Scenario",
        "Test Case",
        "Test Data",
        "Step No.",
        "Step Description",
        "Expected Results",
        "Actual Results",
        "Test Status",
        "Test evidence",
        "Defects",
        "Evidences",
        "Remarks"
    ]
    
    def __init__(self, output_folder: str = "./test_inputs"):
        """
        Initialize Excel writer
        
        Args:
            output_folder: Folder to save Excel files
        """
        self.output_folder = output_folder
        os.makedirs(output_folder, exist_ok=True)
    
    def write_test_cases(self, test_cases: List[TestCase], 
                        filename: str = None) -> str:
        """
        Write test cases to Excel file
        
        Args:
            test_cases: List of TestCase objects
            filename: Output filename (optional)
            
        Returns:
            Path to created Excel file
        """
        if not test_cases:
            logger.warning("No test cases to write")
            return ""
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"test_cases_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_folder, filename)
        
        logger.info(f"Writing {len(test_cases)} test cases to {filepath}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Write headers
        self._write_headers(ws)
        
        # Write test cases
        row_num = 2
        for test_case in test_cases:
            for step_idx, step in enumerate(test_case.steps):
                # First step of a test case includes scenario and case name
                if step_idx == 0:
                    ws[f"A{row_num}"] = test_case.test_scenario
                    ws[f"B{row_num}"] = test_case.test_case_name
                    ws[f"C{row_num}"] = str(test_case.test_data) if test_case.test_data else ""
                    ws[f"L{row_num}"] = test_case.remarks or ""
                else:
                    # Subsequent steps don't repeat scenario/case name
                    ws[f"A{row_num}"] = None
                    ws[f"B{row_num}"] = None
                    ws[f"C{row_num}"] = None
                    ws[f"L{row_num}"] = None
                
                # Step information (always present)
                ws[f"D{row_num}"] = step.step_no
                ws[f"E{row_num}"] = step.description
                ws[f"F{row_num}"] = step.expected_results
                
                # Empty columns for results (filled during execution)
                ws[f"G{row_num}"] = None  # Actual Results
                ws[f"H{row_num}"] = "Pending"  # Test Status
                ws[f"I{row_num}"] = None  # Test evidence
                ws[f"J{row_num}"] = None  # Defects
                ws[f"K{row_num}"] = None  # Evidences
                
                row_num += 1
        
        # Format worksheet
        self._format_worksheet(ws, len(test_cases))
        
        # Save workbook
        try:
            wb.save(filepath)
            logger.info(f"Successfully wrote test cases to {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Error writing Excel file: {str(e)}")
            raise
    
    def write_multiple_files(self, test_cases_by_scenario: dict) -> List[str]:
        """
        Write test cases grouped by scenario to separate Excel files
        
        Args:
            test_cases_by_scenario: Dict mapping scenario name to list of TestCases
            
        Returns:
            List of created file paths
        """
        created_files = []
        
        for scenario_name, test_cases in test_cases_by_scenario.items():
            filename = f"{scenario_name.replace(' ', '_')}.xlsx"
            filepath = self.write_test_cases(test_cases, filename)
            created_files.append(filepath)
        
        logger.info(f"Created {len(created_files)} Excel files")
        return created_files
    
    def append_test_cases(self, filepath: str, new_test_cases: List[TestCase]) -> str:
        """
        Append new test cases to existing Excel file
        
        Args:
            filepath: Path to existing Excel file
            new_test_cases: List of TestCase objects to append
            
        Returns:
            Path to updated Excel file
        """
        try:
            # Load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Find last row
            last_row = ws.max_row
            
            # Append new test cases
            row_num = last_row + 1
            for test_case in new_test_cases:
                for step_idx, step in enumerate(test_case.steps):
                    if step_idx == 0:
                        ws[f"A{row_num}"] = test_case.test_scenario
                        ws[f"B{row_num}"] = test_case.test_case_name
                        ws[f"C{row_num}"] = str(test_case.test_data) if test_case.test_data else ""
                        ws[f"L{row_num}"] = test_case.remarks or ""
                    else:
                        ws[f"A{row_num}"] = None
                        ws[f"B{row_num}"] = None
                        ws[f"C{row_num}"] = None
                        ws[f"L{row_num}"] = None
                    
                    ws[f"D{row_num}"] = step.step_no
                    ws[f"E{row_num}"] = step.description
                    ws[f"F{row_num}"] = step.expected_results
                    ws[f"G{row_num}"] = None
                    ws[f"H{row_num}"] = "Pending"
                    
                    row_num += 1
            
            # Save
            wb.save(filepath)
            logger.info(f"Appended {len(new_test_cases)} test cases to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error appending to Excel file: {str(e)}")
            raise
    
    def _write_headers(self, ws):
        """Write header row with formatting"""
        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            
            # Format header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _format_worksheet(self, ws, test_count: int):
        """Apply formatting to worksheet"""
        # Set column widths
        column_widths = {
            'A': 15,  # Test Scenario
            'B': 20,  # Test Case
            'C': 15,  # Test Data
            'D': 10,  # Step No.
            'E': 35,  # Step Description
            'F': 30,  # Expected Results
            'G': 30,  # Actual Results
            'H': 12,  # Test Status
            'I': 12,  # Test evidence
            'J': 12,  # Defects
            'K': 12,  # Evidences
            'L': 20   # Remarks
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
        
        # Apply cell borders and alignment
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, 
                                min_col=1, max_col=len(self.HEADERS)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Alternate row colors for readability
                if cell.row > 1 and cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", 
                                           end_color="E7E6E6", 
                                           fill_type="solid")
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def create_template(self, filename: str = "test_case_template.xlsx") -> str:
        """
        Create an empty test case template Excel file
        
        Args:
            filename: Template filename
            
        Returns:
            Path to created template file
        """
        logger.info("Creating test case template...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Write headers
        self._write_headers(ws)
        
        # Add sample row
        sample_data = [
            "Login",
            "successful_login",
            None,
            1,
            "Enter valid credentials",
            "User is authenticated",
            None,
            "Pending",
            None,
            None,
            None,
            "Sample test case"
        ]
        
        for col_num, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col_num).value = value
        
        # Format
        self._format_worksheet(ws, 1)
        
        filepath = os.path.join(self.output_folder, filename)
        wb.save(filepath)
        
        logger.info(f"Created template at {filepath}")
        return filepath
