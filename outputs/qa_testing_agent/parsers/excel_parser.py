"""
Excel test case parser
"""
import os
from typing import List, Dict, Any
from openpyxl import load_workbook
from models import TestCase, TestStep
from config.settings import TestStatus
import logging

logger = logging.getLogger(__name__)

class TestCaseParser:
    """Parse test cases from Excel files"""
    
    def __init__(self, input_folder: str = "./test_inputs"):
        self.input_folder = input_folder
        self.test_cases: List[TestCase] = []
    
    def parse_folder(self, file_filter: str | None = None) -> List[TestCase]:
        """Parse Excel files in input folder (optionally filtered by filename)"""
        if not os.path.exists(self.input_folder):
            logger.warning(f"Input folder not found: {self.input_folder}")
            return []

        self.test_cases = []
        
        excel_files = self._resolve_excel_files(file_filter)
        
        logger.info(f"Found {len(excel_files)} Excel files")
        
        for excel_file in excel_files:
            filepath = os.path.join(self.input_folder, excel_file)
            self._parse_excel(filepath)
        
        return self.test_cases

    def list_excel_files(self, file_filter: str | None = None) -> List[str]:
        try:
            return sorted(self._resolve_excel_files(file_filter))
        except Exception as e:
            logger.error(f"Error listing Excel files: {str(e)}")
            return []

    def _resolve_excel_files(self, file_filter: str | None) -> List[str]:
        excel_files = [f for f in os.listdir(self.input_folder)
                       if f.endswith(('.xlsx', '.xls'))]
        if not file_filter:
            return excel_files

        name = str(file_filter).strip()
        if not name:
            return excel_files

        name_lower = name.lower()
        exact = [
            f for f in excel_files
            if f.lower() == name_lower or os.path.splitext(f)[0].lower() == name_lower
        ]
        if exact:
            return exact

        partial = [
            f for f in excel_files
            if name_lower in f.lower() or name_lower in os.path.splitext(f)[0].lower()
        ]
        if not partial:
            logger.warning(f"No Excel files matched filter: {name}")
        return partial
    
    def _parse_excel(self, filepath: str) -> None:
        """Parse a single Excel file"""
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in worksheet[1]]
            header_map = {header: idx for idx, header in enumerate(headers)}
            
            current_scenario = None
            current_case = None
            steps = []
            
            # Parse rows starting from row 2
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), 2):
                row_values = [cell.value for cell in row]
                
                # Check if this is a new test case (has Test Scenario)
                if row_values[header_map.get('Test Scenario', 0)]:
                    # Save previous test case if exists
                    if current_case and steps:
                        current_case.steps = steps
                        self.test_cases.append(current_case)
                    
                    # Start new test case
                    scenario = row_values[header_map.get('Test Scenario', 0)]
                    case_name = row_values[header_map.get('Test Case', 0)]
                    test_data = row_values[header_map.get('Test Data', 0)]
                    remarks = row_values[header_map.get('Remarks', 0)]
                    if isinstance(test_data, str):
                        try:
                            import ast
                            test_data = ast.literal_eval(test_data)
                        except Exception:
                            test_data = {"raw": test_data}
                    if test_data is None:
                        test_data = {}
                    
                    current_scenario = scenario
                    current_case = TestCase(
                        test_scenario=scenario,
                        test_case_name=case_name or f"{scenario}_test",
                        test_data=test_data,
                        remarks=remarks,
                        steps=[]
                    )
                    steps = []
                
                # Parse step if current case exists
                if current_case:
                    step_no = row_values[header_map.get('Step No.', 0)]
                    step_desc = row_values[header_map.get('Step Description', 0)]
                    expected = row_values[header_map.get('Expected Results', 0)]
                    
                    if step_no and step_desc:
                        step = TestStep(
                            step_no=int(step_no) if isinstance(step_no, int) else step_no,
                            description=step_desc,
                            expected_results=expected or "N/A",
                            status=TestStatus.PENDING
                        )
                        steps.append(step)
            
            # Don't forget the last test case
            if current_case and steps:
                current_case.steps = steps
                self.test_cases.append(current_case)
            
            logger.info(f"Parsed {len(self.test_cases)} test cases from {filepath}")
            
        except Exception as e:
            logger.error(f"Error parsing {filepath}: {str(e)}")
            raise
    
    def get_test_cases(self) -> List[TestCase]:
        """Get all parsed test cases"""
        return self.test_cases
